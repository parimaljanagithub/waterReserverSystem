from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app
from flask_login import login_user, current_user, logout_user, login_required 
from src.upload.forms import UploadCsvFileForm
from src import db
from src.model.WaterReserverModel import Dam, DemandType, DemandData, Storage, Inflow
import pandas as pd 
#import xlrd 
import openpyxl
import os
import secrets
from src.main.utils import getDateFromString, getStartDayOfTheMonth, getEndDayOfTheMonth

upload = Blueprint('upload', __name__)

@upload.route("/file/upload", methods=['GET', 'POST'])
@login_required
def uploadFile():
    form = UploadCsvFileForm()
    damlist = db.session.query(Dam)
    demandType = db.session.query(DemandType)

    if form.validate_on_submit():
        if form.file_upload.data:
            selected_file_type = request.form.get('inlineRadioOptions')
            selected_demand_type_id = request.form.get('select_demand_type')
            selected_dam_id = request.form.get('select_dam') 
            
            if selected_file_type == 'DemanFile' :
                file_path = save_csv_to_server(form.file_upload.data,'D_') # save file to static folder
                sheet_obj = read_csv_data_from_server(file_path) # read file from static folder
                status = save_demand_data_to_DataBase(sheet_obj, selected_demand_type_id)
                if status == 'success':
                    flash('Your demand data has been updated successfully', 'success')
            elif selected_file_type == 'StorageFile':
                print('selected_dam_id',selected_dam_id)
                if len(selected_dam_id) == 0:
                    flash('select a particular dam','danger')
                    return render_template('fileUpload.html', title='Data load', form=form, damlist=damlist,\
                                           demandType=demandType)
                file_path = save_csv_to_server(form.file_upload.data, 'S_')
                sheet_obj = read_csv_data_from_server(file_path) # read file from static folder
                status = save_storage_data_to_dataBase(sheet_obj,selected_dam_id)
                if status == 'success':
                    flash('Your Storage data has been updated successfully', 'success')

            elif selected_file_type == 'InflowFile':
                file_path = save_csv_to_server(form.file_upload.data, 'IN_')
                sheet_obj = read_csv_data_from_server(file_path) # read file from static folder
                status = save_inflow_data_to_database(sheet_obj) 
                if status == 'success':
                    flash('Your inflow data has been updated successfully', 'success')
          
        else :
            flash('Select a file to upload!', 'danger')
            return render_template('fileUpload.html', title='Data load', form=form, damlist=damlist, demandType=demandType)
    return render_template('fileUpload.html', title='Data load', form=form, damlist=damlist, demandType=demandType)
    

def save_csv_to_server(from_csv, fileType):
    path =("/Users/parimal.jana/Downloads/biltu_pal_phd/sampleData/putul.xlsx")
    _, f_ext = os.path.splitext(from_csv.filename)
    random_hex = secrets.token_hex(8)
    excel_fn = fileType + random_hex + f_ext
    picture_path = os.path.join(current_app.root_path, 'static/uploaded_files', excel_fn)
    from_csv.save(picture_path)
    print('ggg->', picture_path)
    return picture_path

def read_csv_data_from_server(csv_location):
    #print(csv_location)
    wb_obj = openpyxl.load_workbook(csv_location)
    sheet_obj = wb_obj.active
    return sheet_obj

# This method will store the data for storage table from storage excel
def save_storage_data_to_dataBase(sheet_obj,dam_id):
    total_columns = sheet_obj.max_column
    total_rows = sheet_obj.max_row
    existing_data = db.session.query(Storage).filter_by(dam_id=dam_id).first()
    print('existing_data =', existing_data)
    if existing_data is not None: 
        flash(f'File contain Existing storage data! please upload a file with new data', 'danger')
        return
    storage_data_list = []
    for r in range(2, total_rows+1):
        storage = Storage()
        storage.dam_id = dam_id
        for c in range(1, total_columns+1):
            cell_obj = sheet_obj.cell(row=r, column=c)
            if c == 1:
                storage.storage_value = cell_obj.value
            if c == 2:
                storage.area_value = cell_obj.value
            if c == 3:
                storage.elevation_value = cell_obj.value
        storage_data_list.append(storage)
    db.session.add_all(storage_data_list)
    db.session.commit()
    return 'success'

def save_inflow_data_to_database(sheet_obj):
    # here we are checking the inflow table. if inflow data present then we will not allow 
    # to insert the data
    data = db.session.query(Inflow).first()
    if data is not None:
        return 
        
    total_columns = sheet_obj.max_column
    total_rows = sheet_obj.max_row
    mahi_dam_id = db.session.query(Dam).filter_by(dam_name='MAHI').first().id
    kadana_dam_id = db.session.query(Dam).filter_by(dam_name='KADANA').first().id
    panam_dam_id = db.session.query(Dam).filter_by(dam_name='PANAM').first().id
    print('mashi dam ', mahi_dam_id)
    inflow_data_list = []
    for r in range(2, total_rows+1):
        # mahi near future
        inflow_m = Inflow()
        inflow_m.inflow_data_date = sheet_obj.cell(row=r, column=1).value
        inflow_m.four_near_future = sheet_obj.cell(row=r, column=2).value
        inflow_m.eight_near_future = sheet_obj.cell(row=r, column=3).value
        inflow_m.dam_id = mahi_dam_id
        inflow_data_list.append(inflow_m)
        # kadana near future
        inflow_k = Inflow()
        inflow_k.inflow_data_date = sheet_obj.cell(row=r, column=1).value
        inflow_k.four_near_future = sheet_obj.cell(row=r, column=4).value
        inflow_k.eight_near_future = sheet_obj.cell(row=r, column=5).value
        inflow_k.dam_id = kadana_dam_id
        inflow_data_list.append(inflow_k)
        # panam near future
        inflow_p = Inflow()
        inflow_p.inflow_data_date = sheet_obj.cell(row=r, column=1).value
        inflow_p.four_near_future = sheet_obj.cell(row=r, column=6).value
        inflow_p.eight_near_future = sheet_obj.cell(row=r, column=7).value
        inflow_p.dam_id = kadana_dam_id
        inflow_data_list.append(inflow_p)
         
        # 8th is the blank column

        # mahi far future
        inflow_mf = Inflow()
        inflow_mf.inflow_data_date = sheet_obj.cell(row=r, column=9).value
        inflow_mf.four_far_future = sheet_obj.cell(row=r, column=10).value
        inflow_mf.eight_far_future = sheet_obj.cell(row=r, column=11).value
        inflow_mf.dam_id = mahi_dam_id
        inflow_data_list.append(inflow_mf)
        # kadana far future
        inflow_kf = Inflow()
        inflow_kf.inflow_data_date = sheet_obj.cell(row=r, column=9).value
        inflow_kf.four_far_future = sheet_obj.cell(row=r, column=12).value
        inflow_kf.eight_far_future = sheet_obj.cell(row=r, column=13).value
        inflow_kf.dam_id = kadana_dam_id
        inflow_data_list.append(inflow_kf)
        # panam far future
        inflow_pf = Inflow()
        inflow_pf.inflow_data_date = sheet_obj.cell(row=r, column=9).value
        inflow_pf.four_near_future = sheet_obj.cell(row=r, column=14).value
        inflow_pf.eight_near_future = sheet_obj.cell(row=r, column=15).value
        inflow_pf.dam_id = kadana_dam_id
        inflow_data_list.append(inflow_pf)
        

    print('length of inflow_data_list = ', len(inflow_data_list))
    
    for i in range(0, len(inflow_data_list)):
        print(inflow_data_list[i])

    print('length of inflow_data_list = ', len(inflow_data_list))
    print('total rows', total_rows)
    db.session.add_all(inflow_data_list)
    db.session.commit()
    return 'success'
        
            

                
            
            
           

            
            

            
            
            



def save_demand_data_to_DataBase(sheet_obj,  demand_type_id):
    total_columns = sheet_obj.max_column
    total_rows = sheet_obj.max_row
    print('total rows=',total_rows, 'total column =',total_columns)
    demand_data_list = []
    for r in range(2, total_rows+1): 
        
        date_cell_obj = sheet_obj.cell(row=r, column=1)
        print('row no =', r,'-->', date_cell_obj.value)
        for c in range(2, total_columns+1):
            cell_obj = sheet_obj.cell(row=r, column=c)

            demand_data = DemandData(demand_data_date=date_cell_obj.value, demand_value= cell_obj.value,
                         demand_type_id=demand_type_id)
            if c == 2:
               mahi = db.session.query(Dam).filter_by(dam_name='MAHI').first()
               demand_data.dam_id = mahi.id
            elif c == 3:
                kadana = db.session.query(Dam).filter_by(dam_name='KADANA').first()
                demand_data.dam_id = kadana.id
            elif c == 4:
                panam = db.session.query(Dam).filter_by(dam_name='PANAM').first()
                demand_data.dam_id = panam.id
            
            # check if data is present for that month
            data = db.session.query(DemandData).\
            filter( DemandData.demand_type_id == demand_type_id, DemandData.dam_id== demand_data.dam_id, \
            DemandData.demand_data_date>=getStartDayOfTheMonth(date_cell_obj.value),\
            DemandData.demand_data_date<=getEndDayOfTheMonth(date_cell_obj.value),\
            ).first()

            if data is not None: 
                flash(f'File contain Existing data! please upload a file with new data', 'danger')
                return   # existing from method 
            demand_data_list.append(demand_data)
    # this line mean data is not present in db and we need to insert it 
    db.session.add_all(demand_data_list)
    db.session.commit()
    print('data successfully inserted into db')
    return 'success'
 
    

   
    
    